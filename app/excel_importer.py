from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ALIASES = {
    "sku": {"cod", "sku", "product_code", "default_code", "cod produs", "cod_produs"},
    "title": {"titlu", "title", "name", "nume", "produs", "product_name"},
    "description": {"descriere", "description", "website_description", "detalii"},
    "short_description": {"descriere scurta", "description_sale", "short_description"},
    "price": {"pret", "price", "list_price", "pret vanzare", "pret_vanzare"},
    "quantity": {"cantitate", "quantity", "qty", "stoc"},
    "category_name": {"categorie", "category", "public_category", "categoria"},
    "brand": {"marca", "brand"},
    "image_url": {"imagine", "image", "image_url", "poza", "url poza"},
}


def parse_excel(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip().lower() for cell in rows[0]]
    mapped = [_map_header(h) for h in headers]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        item: dict[str, Any] = {}
        for idx, value in enumerate(row):
            key = mapped[idx] if idx < len(mapped) else None
            if not key or value is None:
                continue
            if key in {"price", "quantity"}:
                item[key] = _float(value)
            elif key == "image_url":
                item.setdefault("image_urls", []).append(str(value).strip())
            else:
                item[key] = str(value).strip()
        if item.get("title") or item.get("sku"):
            result.append(item)
    return result


def _map_header(header: str) -> str | None:
    clean = " ".join(header.replace("_", " ").split()).lower()
    for key, names in ALIASES.items():
        if clean in names:
            return key
    return None


def _float(value: object) -> float:
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0
